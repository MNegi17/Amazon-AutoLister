import openpyxl
from openpyxl.utils import get_column_letter
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
import os

# Register OpenXML namespaces at global module level
ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
ET.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
ET.register_namespace('x14ac', 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac')

NS_MAIN = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

def col_to_index(col_str):
    col_str = col_str.upper()
    exp = 0
    idx = 0
    for char in reversed(col_str):
        idx += (ord(char) - ord('A') + 1) * (26 ** exp)
        exp += 1
    return idx

class ExcelProcessor:
    @staticmethod
    def clean_label(label):
        if not label:
            return ""
        # Remove suffix like " - [ SHIRT ]" or " - [  ]"
        return re.sub(r"\s*-\s*\[.*\]\s*$", "", str(label)).strip()

    @classmethod
    def parse_template(cls, filepath):
        """
        Parses an Amazon category template and returns metadata.
        Returns:
            dict containing:
                - 'attributes': dict of tech_name -> attribute info
                - 'ptd_mappings': dict of product_type -> list of tech_names
                - 'valid_values': dict of label -> list of allowed values
                - 'sheet_info': dict of template indices and layout
        """
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        
        # 1. Parse Valid Values Sheet
        valid_values_map = {}
        if 'Valid Values' in wb.sheetnames:
            sheet_vv = wb['Valid Values']
            for row in sheet_vv.iter_rows(values_only=True):
                if len(row) > 1 and row[1]:
                    raw_label = str(row[1])
                    if " - [" in raw_label:
                        clean_lbl = cls.clean_label(raw_label)
                        allowed = [str(x).strip() for x in row[2:] if x is not None]
                        if allowed:
                            valid_values_map[clean_lbl] = allowed

        # 2. Parse Data Definitions Sheet
        definitions_map = {}
        if 'Data Definitions' in wb.sheetnames:
            sheet_dd = wb['Data Definitions']
            rows_dd = list(sheet_dd.iter_rows(values_only=True))
            header_row_idx = 1
            for idx, r in enumerate(rows_dd[:5]):
                if r and 'Field Name' in r and 'Required?' in r:
                    header_row_idx = idx
                    break
            
            headers = rows_dd[header_row_idx]
            field_name_idx = headers.index('Field Name') if 'Field Name' in headers else 1
            label_idx = headers.index('Local Label Name') if 'Local Label Name' in headers else 2
            desc_idx = headers.index('Accepted Values') if 'Accepted Values' in headers else 3
            required_idx = headers.index('Required?') if 'Required?' in headers else 5
            
            for r in rows_dd[header_row_idx + 1:]:
                if len(r) > field_name_idx and r[field_name_idx]:
                    tech_name = str(r[field_name_idx]).strip()
                    label_name = str(r[label_idx]).strip() if len(r) > label_idx and r[label_idx] else ""
                    required_status = str(r[required_idx]).strip() if len(r) > required_idx and r[required_idx] else "Optional"
                    desc = str(r[desc_idx]).strip() if len(r) > desc_idx and r[desc_idx] else ""
                    
                    definitions_map[tech_name] = {
                        "label": label_name,
                        "required": required_status,
                        "description": desc
                    }

        # 3. Parse Template Sheet Columns
        attributes = {}
        sheet_info = {}
        if 'Template' in wb.sheetnames:
            sheet_t = wb['Template']
            rows_t = list(sheet_t.iter_rows(max_row=10, values_only=True))
            
            label_row_idx = 3
            attr_row_idx = 4
            data_row_idx = 6
            
            if rows_t and rows_t[0] and rows_t[0][0] and "settings=" in str(rows_t[0][0]):
                settings_str = str(rows_t[0][0])
                label_match = re.search(r"labelRow=(\d+)", settings_str)
                attr_match = re.search(r"attributeRow=(\d+)", settings_str)
                data_match = re.search(r"dataRow=(\d+)", settings_str)
                if label_match:
                    label_row_idx = int(label_match.group(1)) - 1
                if attr_match:
                    attr_row_idx = int(attr_match.group(1)) - 1
                if data_match:
                    data_row_idx = int(data_match.group(1)) - 1
            
            sheet_info = {
                "label_row": label_row_idx + 1,
                "attribute_row": attr_row_idx + 1,
                "data_row": data_row_idx + 1
            }
            
            labels = rows_t[label_row_idx]
            tech_names = rows_t[attr_row_idx]
            
            for col_idx in range(len(tech_names)):
                tech_name = tech_names[col_idx]
                if tech_name:
                    tech_name = str(tech_name).strip()
                    label = str(labels[col_idx]).strip() if col_idx < len(labels) and labels[col_idx] else ""
                    col_letter = get_column_letter(col_idx + 1)
                    
                    def_info = definitions_map.get(tech_name, {})
                    required = def_info.get("required", "Optional")
                    
                    clean_lbl = cls.clean_label(label)
                    allowed_values = valid_values_map.get(clean_lbl, None)
                    if not allowed_values:
                        allowed_values = valid_values_map.get(label, None)
                        
                    attributes[tech_name] = {
                        "technical_name": tech_name,
                        "label": label or def_info.get("label", ""),
                        "column_letter": col_letter,
                        "column_index": col_idx,
                        "required": required,
                        "valid_values": allowed_values,
                        "description": def_info.get("description", "")
                    }

        # 4. Parse AttributePTDMAP Sheet
        ptd_mappings = {}
        if 'AttributePTDMAP' in wb.sheetnames:
            sheet_ptd = wb['AttributePTDMAP']
            rows_ptd = list(sheet_ptd.iter_rows(values_only=True))
            if rows_ptd:
                header = rows_ptd[0]
                ptds = []
                for col_idx in range(1, len(header)):
                    if header[col_idx]:
                        ptds.append((col_idx, str(header[col_idx]).strip()))
                
                for r in rows_ptd[1:]:
                    if r and r[0]:
                        attr_name = str(r[0]).strip()
                        for col_idx, ptd_name in ptds:
                            if col_idx < len(r) and (r[col_idx] == 1 or str(r[col_idx]) == '1'):
                                if ptd_name not in ptd_mappings:
                                    ptd_mappings[ptd_name] = []
                                ptd_mappings[ptd_name].append(attr_name)
                                
        wb.close()
        return {
            "attributes": attributes,
            "ptd_mappings": ptd_mappings,
            "valid_values": valid_values_map,
            "sheet_info": sheet_info
        }

    @classmethod
    def write_template(cls, template_path: str, output_path: str, generated_rows: list, sheet_info: dict):
        """
        Copies the Amazon template file and writes generated_rows into the Template sheet
        starting at data_row (Row 7), modifying worksheet XML directly inside the ZIP package.
        Enforces clean OpenXML default namespace elements and strict ascending column sorting (A7, B7, C7...).
        Preserves 100% of macros, data validation extensions, drawings, relationships, and formatting.
        """
        shutil.copy2(template_path, output_path)

        sheet_target_rel = None
        with zipfile.ZipFile(output_path, 'r') as zin:
            if "xl/workbook.xml" in zin.namelist():
                wb_xml = zin.read("xl/workbook.xml")
                wb_root = ET.fromstring(wb_xml)
                
                rels_map = {}
                if "xl/_rels/workbook.xml.rels" in zin.namelist():
                    rels_xml = zin.read("xl/_rels/workbook.xml.rels")
                    rels_root = ET.fromstring(rels_xml)
                    for rel in rels_root.findall('{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                        r_id = rel.get('Id')
                        target = rel.get('Target')
                        if r_id and target:
                            rels_map[r_id] = target

                for sheet in wb_root.findall(f'.//{{{NS_MAIN}}}sheet'):
                    if sheet.get('name') == 'Template':
                        r_id = None
                        for k, v in sheet.attrib.items():
                            if 'id' in k.lower() and k != 'sheetId':
                                r_id = v
                                break
                        target = rels_map.get(r_id)
                        if target:
                            if not target.startswith('xl/'):
                                target = 'xl/' + target
                            sheet_target_rel = target
                        break

        if not sheet_target_rel:
            sheet_target_rel = "xl/worksheets/sheet5.xml"

        with zipfile.ZipFile(output_path, 'r') as zin:
            sheet_bytes = zin.read(sheet_target_rel)
            shared_strings = []
            if "xl/sharedStrings.xml" in zin.namelist():
                ss_bytes = zin.read("xl/sharedStrings.xml")
                ss_root = ET.fromstring(ss_bytes)
                for si in ss_root.findall(f'{{{NS_MAIN}}}si'):
                    t_vals = [t.text for t in si.findall(f'.//{{{NS_MAIN}}}t') if t.text]
                    shared_strings.append("".join(t_vals))

        root = ET.fromstring(sheet_bytes)
        attr_row_num = str(sheet_info.get("attribute_row", 5))
        data_row_start = sheet_info.get("data_row", 7)

        sheetData = root.find(f'{{{NS_MAIN}}}sheetData')
        if sheetData is None:
            raise ValueError("sheetData element not found in Template worksheet XML.")

        attr_row_elem = None
        for r in sheetData.findall(f'{{{NS_MAIN}}}row'):
            if r.get('r') == attr_row_num:
                attr_row_elem = r
                break

        col_map = {}
        if attr_row_elem is not None:
            for c in attr_row_elem.findall(f'{{{NS_MAIN}}}c'):
                cell_ref = c.get('r')
                col_letter = re.sub(r'\d+', '', cell_ref)
                cell_type = c.get('t')
                
                val_text = None
                v_elem = c.find(f'{{{NS_MAIN}}}v')
                val = v_elem.text if v_elem is not None else None
                if cell_type == 's' and val and val.isdigit() and int(val) < len(shared_strings):
                    val_text = shared_strings[int(val)]
                elif val:
                    val_text = val
                else:
                    t_elem = c.find(f'.//{{{NS_MAIN}}}t')
                    if t_elem is not None and t_elem.text:
                        val_text = t_elem.text
                        
                if val_text:
                    col_map[str(val_text).strip()] = col_letter

        def get_or_create_row(r_num_str):
            for r in sheetData.findall(f'{{{NS_MAIN}}}row'):
                if r.get('r') == r_num_str:
                    return r
            r_num_int = int(r_num_str)
            new_r = ET.Element(f'{{{NS_MAIN}}}row', {'r': r_num_str})
            inserted = False
            for idx, child in enumerate(list(sheetData)):
                if child.tag.endswith('row'):
                    curr_r_int = int(child.get('r', '0'))
                    if curr_r_int > r_num_int:
                        sheetData.insert(idx, new_r)
                        inserted = True
                        break
            if not inserted:
                sheetData.append(new_r)
            return new_r

        for row_offset, row_data in enumerate(generated_rows):
            target_r_num = data_row_start + row_offset
            r_str = str(target_r_num)
            row_elem = get_or_create_row(r_str)

            for tech_name, value in row_data.items():
                if tech_name.startswith("__") or value is None:
                    continue
                col_let = col_map.get(tech_name)
                if not col_let:
                    continue

                cell_ref = f"{col_let}{r_str}"
                
                existing_c = None
                for c in row_elem.findall(f'{{{NS_MAIN}}}c'):
                    if c.get('r') == cell_ref:
                        existing_c = c
                        break

                if existing_c is None:
                    existing_c = ET.SubElement(row_elem, f'{{{NS_MAIN}}}c', {'r': cell_ref, 't': 'inlineStr'})
                    is_elem = ET.SubElement(existing_c, f'{{{NS_MAIN}}}is')
                    t_elem = ET.SubElement(is_elem, f'{{{NS_MAIN}}}t')
                    t_elem.text = str(value)
                else:
                    existing_c.set('t', 'inlineStr')
                    for ch in list(existing_c):
                        existing_c.remove(ch)
                    is_elem = ET.SubElement(existing_c, f'{{{NS_MAIN}}}is')
                    t_elem = ET.SubElement(is_elem, f'{{{NS_MAIN}}}t')
                    t_elem.text = str(value)

            # Enforce strict OpenXML ascending column sorting (A, B, C, D...)
            child_cells = list(row_elem.findall(f'{{{NS_MAIN}}}c'))
            if child_cells:
                sorted_cells = sorted(
                    child_cells,
                    key=lambda cell: col_to_index(re.sub(r'\d+', '', cell.get('r', 'A1')))
                )
                for cell in child_cells:
                    row_elem.remove(cell)
                for cell in sorted_cells:
                    row_elem.append(cell)

        new_sheet_xml = ET.tostring(root, encoding='utf-8', xml_declaration=True)

        tmp_zip = output_path + ".tmp.zip"
        with zipfile.ZipFile(output_path, 'r') as zin:
            with zipfile.ZipFile(tmp_zip, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    content = zin.read(item.filename)
                    if item.filename == sheet_target_rel:
                        zout.writestr(item, new_sheet_xml)
                    else:
                        zout.writestr(item, content)

        if os.path.exists(output_path):
            os.remove(output_path)
        shutil.move(tmp_zip, output_path)
